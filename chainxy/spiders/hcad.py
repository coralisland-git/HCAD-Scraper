# from __future__ import unicode_literals
import scrapy

import json

import os

import scrapy

from scrapy.spiders import Spider

from scrapy.http import FormRequest

from scrapy.http import Request

from chainxy.items import ChainItem

from selenium import webdriver

import time

from lxml import etree

from lxml import html

import pdb

from openpyxl import load_workbook

from openpyxl import Workbook 





class hcad(scrapy.Spider):

	name = 'hcad'

	domain = ''

	history = []

	locations = []

	history = []

	def __init__(self):

		script_dir = os.path.dirname(__file__)

		file_path = script_dir + '/data/input.xlsx'

		workbook = load_workbook( file_path )

		sheet = workbook[workbook.sheetnames[0]]

		index = 0

		for record in sheet.rows:


			location = {
				'address' : record[0].value.replace(record[0].value.split(" ")[0], '').strip(),

				'num' : record[0].value.split(" ")[0],

				'zipcode' : record[1].value

			}

			self.locations.append(location)


	def start_requests(self):

		yield scrapy.Request('http://hcad.org/quick-search', callback=self.main)

	
	def main(self, response):

		for location in self.locations[1:]:

			# pdb.set_trace()

			num = location['num']

			address = location['address']

			zipcode = location['zipcode']
				# address="KNOTTY OAKS TRL"
				# zipcode = '5207'

			url = 'https://public.hcad.org/records/QuickRecord.asp'

			header= {			
				'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8',
				'Cache-Control': 'max-age=0',
				'Connection': 'keep-alive',
				'Host': 'public.hcad.org',
				'Origin': 'https://public.hcad.org',
				'Referer': 'https://public.hcad.org/records/QuickSearch.asp',
				'Upgrade-Insecure-Requests': '1',
				'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
				'Accept-Encoding': 'gzip, deflate, br',
				'Content-Type': 'application/x-www-form-urlencoded',
				'Cookie': '_ga=GA1.2.1367539757.1532363385; _gid=GA1.2.1753308311.1532363385; ASPSESSIONIDSUADTBSA=MNIKEOPCAPGKIBGJGPBAMHNH',
				'Upgrade-Insecure-Requests': '1',
				'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.87 Safari/537.36'
			}
			
			formdata = {

				'TaxYear': '2018',

				'stnum': num,

				'stname': address
			}


			yield scrapy.FormRequest(url, callback=self.body, headers=header,  method="POST", formdata=formdata, meta = {
	                      'dont_redirect': True,
	                      'handle_httpstatus_list': [302],
	                      'property_address' : address,
	                      'property_num' : num,
	                  	  'property_zip' : zipcode
	                  }, dont_filter=True )

	
	def body(self, response):

		try:
			url = 'https://public.hcad.org'+response.headers['Location']
			yield scrapy.Request(url, callback=self.parse_page, meta={
	              	'property_address' : response.meta['property_address'],
	              	'property_num' : response.meta['property_num'],
	              	'property_zip' : response.meta['property_zip']
	              })
		except:
			pass

	def parse_page(self, response):

		data = self.eliminate_space(response.xpath('//table[@class="bgcolor_1"][5]/tr[2]/td[1]//th//text()').extract())

		item = ChainItem()

		try:

			item['property_address'] = response.meta['property_num'] + ' ' + response.meta['property_address']

			item['propert_zip'] = response.meta['property_zip']

			item['owner_name'] = data[0]

			begin = len(data)

			item['mailing_address'] = data[begin-2].strip()

			item['mailing_city'] = data[begin-1].split('=')[0]		

			item['mailing_zip'] = data[begin-1].split('=')[2]

			item['mailing_state'] = data[begin-1].split('=')[1]

			item['phone_number'] = ''

			item['phone_number2'] = ''

			if item['property_address'] not in self.history:
				
				self.history.append(item['property_address'])
				
				yield item
		except :

			pass



	def validate(self, item):

		try:

			return item.replace('\n', '').replace('\t','').replace('\r', '').encode('raw-unicode-escape').replace('\xa0', '=').strip()

		except:

			pass


	def eliminate_space(self, items):

	    tmp = []

	    for item in items:

	        if self.validate(item) != '':

	            tmp.append(self.validate(item))

	    return tmp